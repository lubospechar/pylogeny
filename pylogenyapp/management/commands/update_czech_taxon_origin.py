from io import BytesIO
from urllib.request import urlopen

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from pylogenyapp.models import CzechTaxonOrigin, Taxon


class Command(BaseCommand):
    help = "Update Czech taxon origin for taxa from Pladias XLSX file."

    SOURCE_URL = (
        "https://files.ibot.cas.cz/cevs/downloads/"
        "Pladias-taxony-puvod-invazni-status-cerveny-seznam-ochrana-2023-11-06.xlsx"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Only show what would be changed, without saving.",
        )
        parser.add_argument(
            "--url",
            default=self.SOURCE_URL,
            help="Source XLSX URL.",
        )

    def handle(self, *args, **options):
        dry_run = options["dry_run"]
        url = options["url"]

        self.stdout.write(f"Downloading source file from: {url}")

        with urlopen(url) as response:
            source_data = response.read()

        workbook = load_workbook(
            filename=BytesIO(source_data),
            read_only=True,
            data_only=True,
        )
        worksheet = workbook.active

        origins_by_name = {
            origin.origin.strip(): origin
            for origin in CzechTaxonOrigin.objects.all()
        }

        taxa_by_scientific_name = {
            taxon.scientific_name.strip(): taxon
            for taxon in Taxon.objects.select_related(
                "czech_taxon_origin",
            )
        }

        updated_count = 0
        unchanged_count = 0
        skipped_empty_count = 0
        taxon_not_found_count = 0
        origin_not_found_count = 0

        for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    min_col=1,
                    max_col=5,
                    values_only=True,
                ),
                start=2,
        ):
            scientific_name = row[0]
            origin_name = row[4]

            if scientific_name is None or origin_name is None:
                skipped_empty_count += 1
                continue

            scientific_name = str(scientific_name).strip()
            origin_name = str(origin_name).strip()

            if not scientific_name or not origin_name:
                skipped_empty_count += 1
                continue

            taxon = taxa_by_scientific_name.get(scientific_name)

            if taxon is None:
                taxon_not_found_count += 1
                continue

            origin = origins_by_name.get(origin_name)

            if origin is None:
                origin_not_found_count += 1
                self.stdout.write(
                    self.style.WARNING(
                        f"Row {row_number}: CzechTaxonOrigin not found: "
                        f"{origin_name}"
                    )
                )
                continue

            if taxon.czech_taxon_origin_id == origin.id:
                unchanged_count += 1
                continue

            old_value = (
                taxon.czech_taxon_origin.origin
                if taxon.czech_taxon_origin
                else "-"
            )

            self.stdout.write(
                f"Row {row_number}: {taxon.scientific_name}: "
                f"{old_value} -> {origin.origin}"
            )

            if not dry_run:
                taxon.czech_taxon_origin = origin
                taxon.save(
                    update_fields=[
                        "czech_taxon_origin",
                    ]
                )

            updated_count += 1

        workbook.close()

        if dry_run:
            self.stdout.write(
                self.style.WARNING("Dry run only. No changes were saved.")
            )

        self.stdout.write(
            self.style.SUCCESS(f"Updated taxa: {updated_count}")
        )
        self.stdout.write(f"Unchanged taxa: {unchanged_count}")
        self.stdout.write(f"Skipped empty rows: {skipped_empty_count}")
        self.stdout.write(f"Taxa not found in database: {taxon_not_found_count}")
        self.stdout.write(
            f"CzechTaxonOrigin values not found in database: {origin_not_found_count}"
        )