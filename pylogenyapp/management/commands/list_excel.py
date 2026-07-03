from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class Command(BaseCommand):
    help = "List unique values from selected Excel column starting from row 2."

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_file",
            help="Path to Excel file.",
        )
        parser.add_argument(
            "--column",
            default="B",
            help="Excel column to read, for example B or 2.",
        )

    def get_column_index(self, column):
        column = str(column).strip()

        if column.isdigit():
            column_index = int(column)
        else:
            try:
                column_index = column_index_from_string(column.upper())
            except ValueError as error:
                raise CommandError(
                    f"Invalid column: {column}"
                ) from error

        if column_index < 1:
            raise CommandError(
                f"Invalid column: {column}"
            )

        return column_index

    def handle(self, *args, **options):
        excel_file = options["excel_file"]
        column = options["column"]
        column_index = self.get_column_index(column)

        workbook = load_workbook(
            filename=excel_file,
            read_only=True,
            data_only=True,
        )
        worksheet = workbook.active

        values = set()

        for row in worksheet.iter_rows(
                min_row=2,
                min_col=column_index,
                max_col=column_index,
                values_only=True,
        ):
            value = row[0]

            if value is None:
                continue

            value = str(value).strip()

            if not value:
                continue

            values.add(value)

        workbook.close()

        for value in sorted(values):
            self.stdout.write(value)

        self.stdout.write(
            self.style.SUCCESS(
                f"Found {len(values)} unique values in column {column}."
            )
        )