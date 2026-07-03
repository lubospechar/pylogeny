from io import BytesIO
from urllib.request import urlopen

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from pylogenyapp.models import InvasiveStatus, Taxon


class Command(BaseCommand):
    help = "Update invasive status for taxa from Pladias XLSX file."

    SOURCE_URL = (
        "https://files.ibot.cas.cz/cevs/downloads/"
        "Pladias-taxony-puvod-invazni-status-cerveny-seznam-ochrana-2023-11-06.xlsx"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Only show what would be changed, without saving.",
        )
        parser.add_argument(
            "--url",
            default=self.SOURCE_URL,
            help="Source XLSX URL.",
        )

    def handle(self, *args, **options):
        dry_run = options["dry_run"]
        url = options["url"]

        self.stdout.write(f"Downloading source file from: {url}")

        with urlopen(url) as response:
            source_data = response.read()

        workbook = load_workbook(
            filename=BytesIO(source_data),
            read_only=True,
            data_only=True,
        )
        worksheet = workbook.active

        statuses_by_name = {
            status.status.strip(): status
            for status in InvasiveStatus.objects.all()
        }

        taxa_by_scientific_name = {
            taxon.scientific_name.strip(): taxon
            for taxon in Taxon.objects.select_related(
                "invasive_status",
            )
        }

        updated_count = 0
        unchanged_count = 0
        skipped_empty_count = 0
        taxon_not_found_count = 0
        status_not_found_count = 0

        for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    min_col=1,
                    max_col=6,
                    values_only=True,
                ),
                start=2,
        ):
            scientific_name = row[0]
            invasive_status_name = row[5]

            if scientific_name is None or invasive_status_name is None:
                skipped_empty_count += 1
                continue

            scientific_name = str(scientific_name).strip()
            invasive_status_name = str(invasive_status_name).strip()

            if not scientific_name or not invasive_status_name:
                skipped_empty_count += 1
                continue

            taxon = taxa_by_scientific_name.get(scientific_name)

            if taxon is None:
                taxon_not_found_count += 1
                continue

            invasive_status = statuses_by_name.get(invasive_status_name)

            if invasive_status is None:
                status_not_found_count += 1
                self.stdout.write(
                    self.style.WARNING(
                        f"Row {row_number}: InvasiveStatus not found: "
                        f"{invasive_status_name}"
                    )
                )
                continue

            if taxon.invasive_status_id == invasive_status.id:
                unchanged_count += 1
                continue

            old_value = (
                taxon.invasive_status.status
                if taxon.invasive_status
                else "-"
            )

            self.stdout.write(
                f"Row {row_number}: {taxon.scientific_name}: "
                f"{old_value} -> {invasive_status.status}"
            )

            if not dry_run:
                taxon.invasive_status = invasive_status
                taxon.save(
                    update_fields=[
                        "invasive_status",
                    ]
                )

            updated_count += 1

        workbook.close()

        if dry_run:
            self.stdout.write(
                self.style.WARNING("Dry run only. No changes were saved.")
            )

        self.stdout.write(
            self.style.SUCCESS(f"Updated taxa: {updated_count}")
        )
        self.stdout.write(f"Unchanged taxa: {unchanged_count}")
        self.stdout.write(f"Skipped empty rows: {skipped_empty_count}")
        self.stdout.write(f"Taxa not found in database: {taxon_not_found_count}")
        self.stdout.write(
            f"InvasiveStatus values not found in database: {status_not_found_count}"
        )