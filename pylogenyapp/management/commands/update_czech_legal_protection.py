from io import BytesIO
from urllib.request import urlopen

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from pylogenyapp.models import CzechLegalProtection, Taxon


class Command(BaseCommand):
    help = "Update Czech legal protection for taxa from Pladias XLSX file."

    SOURCE_URL = (
        "https://pladias.ibot.cas.cz/public/traits/downloadTraitData/"
        "feature/218/lang/cs"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Only show what would be changed, without saving.",
        )
        parser.add_argument(
            "--url",
            default=self.SOURCE_URL,
            help="Source XLSX URL.",
        )

    def handle(self, *args, **options):
        dry_run = options["dry_run"]
        url = options["url"]

        self.stdout.write(f"Downloading source file from: {url}")

        with urlopen(url) as response:
            source_data = response.read()

        workbook = load_workbook(
            filename=BytesIO(source_data),
            read_only=True,
            data_only=True,
        )
        worksheet = workbook.active

        legal_protections_by_description = {
            legal_protection.description.strip(): legal_protection
            for legal_protection in CzechLegalProtection.objects.all()
        }

        taxa_by_scientific_name = {
            taxon.scientific_name.strip(): taxon
            for taxon in Taxon.objects.select_related(
                "czech_legal_protection",
            )
        }

        updated_count = 0
        unchanged_count = 0
        skipped_empty_count = 0
        taxon_not_found_count = 0
        legal_protection_not_found_count = 0

        for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    min_col=1,
                    max_col=2,
                    values_only=True,
                ),
                start=2,
        ):
            scientific_name = row[0]
            legal_protection_description = row[1]

            if scientific_name is None or legal_protection_description is None:
                skipped_empty_count += 1
                continue

            scientific_name = str(scientific_name).strip()
            legal_protection_description = str(
                legal_protection_description,
            ).strip()

            if not scientific_name or not legal_protection_description:
                skipped_empty_count += 1
                continue

            taxon = taxa_by_scientific_name.get(scientific_name)

            if taxon is None:
                taxon_not_found_count += 1
                continue

            legal_protection = legal_protections_by_description.get(
                legal_protection_description,
            )

            if legal_protection is None:
                legal_protection_not_found_count += 1
                self.stdout.write(
                    self.style.WARNING(
                        f"Row {row_number}: CzechLegalProtection description "
                        f"not found: {legal_protection_description}"
                    )
                )
                continue

            if taxon.czech_legal_protection_id == legal_protection.id:
                unchanged_count += 1
                continue

            old_value = (
                taxon.czech_legal_protection.description
                if taxon.czech_legal_protection
                else "-"
            )

            self.stdout.write(
                f"Row {row_number}: {taxon.scientific_name}: "
                f"{old_value} -> {legal_protection.description}"
            )

            if not dry_run:
                taxon.czech_legal_protection = legal_protection
                taxon.save(
                    update_fields=[
                        "czech_legal_protection",
                    ]
                )

            updated_count += 1

        workbook.close()

        if dry_run:
            self.stdout.write(
                self.style.WARNING("Dry run only. No changes were saved.")
            )

        self.stdout.write(
            self.style.SUCCESS(f"Updated taxa: {updated_count}")
        )
        self.stdout.write(f"Unchanged taxa: {unchanged_count}")
        self.stdout.write(f"Skipped empty rows: {skipped_empty_count}")
        self.stdout.write(f"Taxa not found in database: {taxon_not_found_count}")
        self.stdout.write(
            "CzechLegalProtection descriptions not found in database: "
            f"{legal_protection_not_found_count}"
        )