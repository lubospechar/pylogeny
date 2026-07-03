from io import BytesIO
from urllib.request import urlopen

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from pylogenyapp.models import EllenbergIndicatorValue, Taxon


class Command(BaseCommand):
    help = "Update Ellenberg indicator values for taxa from XLSX file."

    SOURCE_URL = (
        "https://files.ibot.cas.cz/cevs/downloads/"
        "ekologicke_indikacni_hodnoty.xlsx"
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Only show what would be changed, without saving.",
        )
        parser.add_argument(
            "--url",
            default=self.SOURCE_URL,
            help="Source XLSX URL.",
        )

    def clean_indicator_value(self, value):
        if value is None:
            return None

        value = str(value).strip()

        if not value:
            return None

        if value.endswith(".0"):
            value = value[:-2]

        return int(value)

    def clean_scientific_name(self, value):
        if value is None:
            return ""

        return str(value).strip()

    def format_indicator_values(self, values):
        return (
            f"L: {values['light']}, "
            f"T: {values['temperature']}, "
            f"M: {values['moisture']}, "
            f"R: {values['reaction']}, "
            f"N: {values['nutrients']}, "
            f"S: {values['salinity']}"
        )

    def handle(self, *args, **options):
        dry_run = options["dry_run"]
        url = options["url"]

        self.stdout.write(f"Downloading source file from: {url}")

        with urlopen(url) as response:
            source_data = response.read()

        workbook = load_workbook(
            filename=BytesIO(source_data),
            read_only=True,
            data_only=True,
        )
        worksheet = workbook.active

        taxa_by_scientific_name = {
            taxon.scientific_name.strip(): taxon
            for taxon in Taxon.objects.select_related(
                "ellenberg_indicator_values",
            )
        }

        updated_count = 0
        created_count = 0
        unchanged_count = 0
        skipped_empty_count = 0
        taxon_not_found_count = 0
        invalid_value_count = 0

        for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    min_col=1,
                    max_col=11,
                    values_only=True,
                ),
                start=2,
        ):
            scientific_name = self.clean_scientific_name(row[0])

            try:
                indicator_values_data = {
                    "light": self.clean_indicator_value(row[5]),
                    "temperature": self.clean_indicator_value(row[6]),
                    "moisture": self.clean_indicator_value(row[7]),
                    "reaction": self.clean_indicator_value(row[8]),
                    "nutrients": self.clean_indicator_value(row[9]),
                    "salinity": self.clean_indicator_value(row[10]),
                }
            except ValueError:
                invalid_value_count += 1
                self.stdout.write(
                    self.style.WARNING(
                        f"Row {row_number}: Invalid Ellenberg value."
                    )
                )
                continue

            if not scientific_name:
                skipped_empty_count += 1
                continue

            if not any(value is not None for value in indicator_values_data.values()):
                skipped_empty_count += 1
                continue

            taxon = taxa_by_scientific_name.get(scientific_name)

            if taxon is None:
                taxon_not_found_count += 1
                continue

            indicator_values = taxon.ellenberg_indicator_values

            if indicator_values is None:
                self.stdout.write(
                    f"Row {row_number}: {taxon.scientific_name}: "
                    f"create {self.format_indicator_values(indicator_values_data)}"
                )

                if not dry_run:
                    indicator_values = EllenbergIndicatorValue.objects.create(
                        **indicator_values_data,
                    )
                    taxon.ellenberg_indicator_values = indicator_values
                    taxon.save(
                        update_fields=[
                            "ellenberg_indicator_values",
                        ]
                    )

                created_count += 1
                continue

            has_changed = any(
                getattr(indicator_values, field) != value
                for field, value in indicator_values_data.items()
            )

            if not has_changed:
                unchanged_count += 1
                continue

            self.stdout.write(
                f"Row {row_number}: {taxon.scientific_name}: "
                f"{indicator_values} -> "
                f"{self.format_indicator_values(indicator_values_data)}"
            )

            if not dry_run:
                for field, value in indicator_values_data.items():
                    setattr(indicator_values, field, value)

                indicator_values.save(
                    update_fields=list(indicator_values_data.keys())
                )

            updated_count += 1

        workbook.close()

        if dry_run:
            self.stdout.write(
                self.style.WARNING("Dry run only. No changes were saved.")
            )

        self.stdout.write(
            self.style.SUCCESS(f"Created Ellenberg values: {created_count}")
        )
        self.stdout.write(
            self.style.SUCCESS(f"Updated Ellenberg values: {updated_count}")
        )
        self.stdout.write(f"Unchanged Ellenberg values: {unchanged_count}")
        self.stdout.write(f"Skipped empty rows: {skipped_empty_count}")
        self.stdout.write(f"Invalid value rows: {invalid_value_count}")
        self.stdout.write(f"Taxa not found in database: {taxon_not_found_count}")